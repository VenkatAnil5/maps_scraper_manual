import json
import pandas as pd
import streamlit as st
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import streamlit as st
import pandas as pd
import re
from io import StringIO


# Load the data from the 'data.json' file
with open('data.json', 'r') as file:
    location_data = json.load(file)

# Function to generate dropdown options from the 'data.json' file
def generate_location_dropdown(location_data):
    """Generates location dropdowns based on 'data.json'"""
    state_options = [state['state'] for state in location_data]
    district_options = []
    sub_district_options = []
    village_options = []

    # Function to get districts based on state selection
    def get_districts(state):
        for s in location_data:
            if s['state'] == state:
                return [district['district'] for district in s['districts']]
        return []

    # Function to get sub-districts based on district selection
    def get_sub_districts(state, district):
        for s in location_data:
            if s['state'] == state:
                for d in s['districts']:
                    if d['district'] == district:
                        return [sub['subDistrict'] for sub in d['subDistricts']]
        return []

    # Function to get villages based on sub-district selection
    def get_villages(state, district, sub_district):
        for s in location_data:
            if s['state'] == state:
                for d in s['districts']:
                    if d['district'] == district:
                        for sub in d['subDistricts']:
                            if sub['subDistrict'] == sub_district:
                                return sub['villages']
        return []

    return state_options, get_districts, get_sub_districts, get_villages

# Streamlit Web App
# Streamlit Web App
st.title('Google Maps Scraping')

# Dropdowns for category and location
category = st.selectbox('Select Category', ['Nurseries', 'Landscapers', 'Construction Companies', 'Architects', 'Event Managers'])
state_options, get_districts, get_sub_districts, get_villages = generate_location_dropdown(location_data)

# Select state, district, sub-district, and village
selected_state = st.selectbox('Select State', state_options)
selected_district = st.selectbox('Select District', get_districts(selected_state))
selected_sub_district = st.selectbox('Select Sub-District', get_sub_districts(selected_state, selected_district))
selected_village = st.selectbox('Select Village (optional)', get_villages(selected_state, selected_district, selected_sub_district), index=0)

# Combine location query
location_query = f"{selected_village if selected_village else selected_sub_district}, {selected_district}, {selected_state}"

search_query = f"{category} in {location_query}"
maps_url = f"https://www.google.com/maps/search/{search_query.replace(' ', '+')}"

# Manual Search button functionality
if st.button('Manual Search'):
    st.markdown(f"[Click here to search manually on Google Maps]({maps_url})", unsafe_allow_html=True)


import pandas as pd
import openpyxl
import re
from io import BytesIO  # Import BytesIO
import streamlit as st


def parse_landscaping_data(input_text, selected_sub_district=""):
    # Initialize lists to store parsed data
    names = []
    roles = []
    addresses = []
    timings = []
    phone_numbers = []
    links = []  # To store the generated links

    # Split the input into lines for processing
    lines = input_text.splitlines()

    for i in range(len(lines)):
        line = lines[i].strip()

        # Detect name (single line before review/ratings format)
        if i + 1 < len(lines) and (re.match(r"\d+\.\d+\(\d+\)", lines[i + 1].strip()) or lines[i + 1].strip() == "No reviews"):
            names.append(line)

            # Detect role and address
            if i + 2 < len(lines):
                role_address_line = lines[i + 2].strip()
                if "\u00b7" in role_address_line:  # Check for · symbol
                    role, address = map(str.strip, role_address_line.split("\u00b7", 1))
                    roles.append(role)
                    addresses.append(address)
                else:
                    roles.append("Unknown")
                    addresses.append("Unknown")

            # Detect timings
            if i + 3 < len(lines):
                timing_line = lines[i + 3].strip()
                timing_match = re.match(r"(Closed.*?\u00b7|Opens soon.*?\u00b7|Open.*?\u00b7|Temporarily closed.*?\u00b7)", timing_line)
                if timing_match:
                    timings.append(timing_match.group().strip())
                else:
                    timings.append("Unknown")

            # Detect phone number
            phone_number = "Unknown"
            if i + 3 < len(lines):
                phone_line = lines[i + 3].strip()
                phone_match = re.search(r"(?:\+91\s?|\+91-?|0)?\d{5}\s?\d{5}", phone_line)
                if phone_match:
                    phone_number = phone_match.group().strip()
            phone_numbers.append(phone_number)

            search_query = f"{line} in {selected_sub_district}"
            link = f"https://www.google.com/maps/search/{search_query.replace(' ', '+')}"
            links.append(link)

    # Create a DataFrame with parsed data
    data = {
        "Name": names,
        "Role": roles,
        "Address": addresses,
        "Timings": timings,
        "Phone Number": phone_numbers,
        "Link": links  # Add the new "Link" column
    }

    return pd.DataFrame(data)

# Streamlit app
st.title("Landscaping Data Parser")
st.write("Provide your landscaping data in the input box below:")

# Input box for text data
input_text = st.text_area("Enter the landscaping data here:", height=300)

if st.button("Parse Data"):
    if input_text.strip():
        # Parse the input text
        result_df = parse_landscaping_data(input_text)

        # Display the results in a table
        st.write("### Parsed Data")
        st.dataframe(result_df)

        # Export to Excel with hyperlinks
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            result_df.to_excel(writer, index=False, sheet_name="Parsed Data")

            # Add hyperlinks
            workbook = writer.book
            worksheet = writer.sheets["Parsed Data"]
            for row_num, link in enumerate(result_df["Link"], start=2):  # Start from row 2
                worksheet.cell(row=row_num, column=6).value = link
                worksheet.cell(row=row_num, column=6).hyperlink = link
                worksheet.cell(row=row_num, column=6).style = "Hyperlink"

        file_name = f"{search_query.replace(' ', '_')}"
        file_name = f"{file_name.replace(',', '')}"

        st.download_button(
            label="Download Data as Excel",
            data=excel_buffer.getvalue(),
            file_name=f"{file_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.error("Please enter some data to parse.")
